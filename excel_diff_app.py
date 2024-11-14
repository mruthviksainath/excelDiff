import streamlit as st
import pandas as pd
from openpyxl import load_workbook

def compare_excel_sheets(original_excel_path, user_excel_path):
    # Load both Excel files
    original_wb = load_workbook(original_excel_path, data_only=True)
    user_wb = load_workbook(user_excel_path, data_only=True)
    
    # Initialize lists to store different types of changes
    boolean_changes = []
    other_changes = []

    # Compare each sheet in both workbooks
    for sheet_name in original_wb.sheetnames:
        if sheet_name in user_wb.sheetnames:
            original_sheet = original_wb[sheet_name]
            user_sheet = user_wb[sheet_name]

            # Convert sheets to DataFrames for easy comparison
            original_df = pd.DataFrame(original_sheet.values)
            user_df = pd.DataFrame(user_sheet.values)

            # Ensure both DataFrames have the same structure
            max_rows = max(original_df.shape[0], user_df.shape[0])
            max_cols = max(original_df.shape[1], user_df.shape[1])
            original_df = original_df.reindex(index=range(max_rows), columns=range(max_cols))
            user_df = user_df.reindex(index=range(max_rows), columns=range(max_cols))

            # Compare row by row and column by column
            for row_idx in range(max_rows):
                original_row = original_df.iloc[row_idx].fillna("").tolist()
                user_row = user_df.iloc[row_idx].fillna("").tolist()
                
                if original_row != user_row:
                    row_diff = {
                        "Sheet": sheet_name,
                        "Row": row_idx + 1,
                        "Old": original_row,
                        "New": user_row,
                    }

                    # Check if changes are only true/false or other types
                    if all(isinstance(old, bool) and isinstance(new, bool) and old != new 
                           for old, new in zip(original_row, user_row) if old != new):
                        boolean_changes.append(row_diff)
                    else:
                        other_changes.append(row_diff)

    return boolean_changes, other_changes

def highlight_row_differences(old_row, new_row):
    """Return a highlighted new row where differences are marked in green."""
    highlighted_row = []
    for old_val, new_val in zip(old_row, new_row):
        if old_val != new_val:
            highlighted_row.append(f"<span style='color: green; font-weight: bold;'>{new_val}</span>")
        else:
            highlighted_row.append(str(new_val))
    return highlighted_row

# Streamlit interface
st.title("Excel Sheet Difference Checker")
st.write("Upload the original Excel file and the user-updated Excel file to compare differences.")

# Upload file inputs
original_file = st.file_uploader("Upload Original Excel File", type=["xlsx"])
user_file = st.file_uploader("Upload User Excel File", type=["xlsx"])

# Run comparison if both files are uploaded
if original_file and user_file:
    st.write("Comparing files...")

    # Compare the Excel files and retrieve the differences
    boolean_changes, other_changes = compare_excel_sheets(original_file, user_file)

    # Display results
    def display_differences(differences, change_type):
        change_count = len(differences)
        st.write(f"### {change_type} Changes ({change_count} changes found)")
        for diff in differences:
            st.write(f"**Sheet:** {diff['Sheet']} - **Row:** {diff['Row']}")

            # Display rows in a two-row format
            df_display = pd.DataFrame([diff["Old"], highlight_row_differences(diff["Old"], diff["New"])], index=["Old Row", "New Row"])
            st.write(df_display.to_html(escape=False), unsafe_allow_html=True)

            st.write("---")
    
    # Display Boolean changes with count
    display_differences(boolean_changes, "Boolean")

    # Display Other changes with count
    display_differences(other_changes, "Other")
